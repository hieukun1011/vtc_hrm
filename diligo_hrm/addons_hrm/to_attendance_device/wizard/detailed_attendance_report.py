import base64
import logging
from calendar import monthrange
from datetime import date, datetime
from io import BytesIO

import pytz
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

from odoo import fields, models

_logger = logging.getLogger(__name__)

days = ['T2', 'T3', 'T4', 'T5', 'T6', 'T7', 'CN']


class ReportAttendanceDetailed(models.TransientModel):
    _name = 'report.attendance.detailed'

    type = fields.Selection([('area', 'Area'), ('employee', 'Employee')], string='Type',
                            default="area")
    date_start = fields.Date(string='Date From', required=True,
                             default=lambda self: fields.Date.to_string(
                                 date.today().replace(month=date.today().month - 1, day=1)))
    date_end = fields.Date(string='Date To', required=True,
                           default=lambda self: fields.Date.to_string(
                               (datetime.now() + relativedelta(day=1, days=-1)).date()))
    employee_id = fields.Many2one('hr.employee', string='Employee')
    area = fields.Many2one('hr.channel', string='Area')


    def report_attendance_detailed(self):
        simple_pharmacy_provider_attachment = self.env['ir.attachment'].sudo().browse(
            self.env.ref('to_attendance_device.report_attendance_template').id)
        decode = base64.b64decode(simple_pharmacy_provider_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        # key_col_list_date = []
        # key_col_list_sort = []
        # for rec in range(1, (monthrange(self.date_end.year, self.date_end.month)[1]) + 1):
        #     key_col_list_date.append(rec)
        #     sort = date(self.date_end.year, self.date_end.month, rec)
        #     wd = date.weekday(sort)
        #     key_col_list_sort.append(days[wd])
        # print(key_col_list_date)
        # print(key_col_list_sort)
        if self.type == 'area':
            wb.remove_sheet(wb.active)
            # print(self.area.member_ids)
            # attendance = self.env['hr.attendance'].sudo().search([('check_in', '>=', self.date_start),
            #                                                       ('check_in', '<=', self.date_end),
            #                                                       ('employee_id', 'in', self.area.member_ids.ids),
            #                                                       ])
            # print(attendance)

            for record in self.area.member_ids:
                tz = pytz.timezone(record.tz)
                rec_data = []
                ws = wb.create_sheet(record.name)
                attendance = self.env['hr.attendance'].sudo().search([('check_in', '>=', self.date_start),
                                                                      ('check_in', '<=', self.date_end),
                                                                      ('employee_id', '=', record.id),
                                                                      ], order='check_in asc',)
                thin = borders.Side(style='thin')
                all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
                line_font = Font(name='Times New Roman', size=12)
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=16)
                ws['A1'].value = 'BẢNG CHI TIẾT CHẤM CÔNG'
                ws['A1'].font = line_font
                ws['A1'].border = all_border_thin
                ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=16)
                ws['A2'].value = 'Mã nhân viên:' + record.employee_id + '   Tên nhân viên:' + record.name + '  Bộ phận:' + record.channel_id.name
                ws['A2'].font = line_font
                ws['A2'].border = all_border_thin
                ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
                ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
                ws['A3'].value = 'Ngày thường'
                ws['A3'].font = line_font
                ws['A3'].border = all_border_thin
                ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
                ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=2)
                ws['A4'].value = 'Cuối tuần'
                ws['A4'].font = line_font
                ws['A4'].border = all_border_thin
                ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=2)
                ws['A5'].value = 'TỔNG'
                ws['A5'].font = line_font
                ws['A5'].border = all_border_thin
                ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
                ws.merge_cells(start_row=3, start_column=5, end_row=5, end_column=5)
                ws.merge_cells(start_row=3, start_column=6, end_row=3, end_column=7)
                ws['F3'].value = 'Tăng ca 1'
                ws['F3'].font = line_font
                ws['F3'].border = all_border_thin
                ws['F3'].alignment = Alignment(horizontal='center', vertical='center')
                ws.merge_cells(start_row=4, start_column=6, end_row=4, end_column=7)
                ws['F4'].value = 'Tăng ca 2'
                ws['F4'].font = line_font
                ws['F4'].border = all_border_thin
                ws['F4'].alignment = Alignment(horizontal='center', vertical='center')
                ws.merge_cells(start_row=5, start_column=6, end_row=5, end_column=7)
                ws['F5'].value = 'Tăng ca 3'
                ws['F5'].font = line_font
                ws['F5'].border = all_border_thin
                ws['F5'].alignment = Alignment(horizontal='center', vertical='center')
                ws.merge_cells(start_row=3, start_column=9, end_row=5, end_column=9)
                ws.merge_cells(start_row=3, start_column=10, end_row=3, end_column=12)
                ws['J3'].value = 'Đi trễ'
                ws['J3'].font = line_font
                ws['J3'].border = all_border_thin
                ws['J3'].alignment = Alignment(horizontal='center', vertical='center')
                ws.merge_cells(start_row=4, start_column=10, end_row=4, end_column=11)
                ws['J4'].value = 'Số lần'
                ws['J4'].font = line_font
                ws['J4'].border = all_border_thin
                ws['J4'].alignment = Alignment(horizontal='center', vertical='center')
                ws.merge_cells(start_row=5, start_column=10, end_row=5, end_column=11)
                ws['J5'].value = 'Số phút'
                ws['J5'].font = line_font
                ws['J5'].border = all_border_thin
                ws['J5'].alignment = Alignment(horizontal='center', vertical='center')
                ws.merge_cells(start_row=3, start_column=14, end_row=3, end_column=16)
                ws['N3'].value = 'Về sớm'
                ws['N3'].font = line_font
                ws['N3'].border = all_border_thin
                ws['N3'].alignment = Alignment(horizontal='center', vertical='center')
                ws.merge_cells(start_row=4, start_column=14, end_row=4, end_column=15)
                ws['N4'].value = 'Số lần'
                ws['N4'].font = line_font
                ws['N4'].border = all_border_thin
                ws['N4'].alignment = Alignment(horizontal='center', vertical='center')
                ws.merge_cells(start_row=5, start_column=14, end_row=5, end_column=15)
                ws['N5'].value = 'Số phút'
                ws['N5'].font = line_font
                ws['N5'].border = all_border_thin
                ws['N5'].alignment = Alignment(horizontal='center', vertical='center')
                ws.merge_cells(start_row=6, start_column=1, end_row=8, end_column=1)
                ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=15)
                ws['B7'].value = 'V'
                ws['B7'].font = line_font
                ws['B7'].border = all_border_thin
                ws['B7'].alignment = Alignment(horizontal='center', vertical='center')
                ws['C7'].value = 'OM'
                ws['C7'].font = line_font
                ws['C7'].border = all_border_thin
                ws['C7'].alignment = Alignment(horizontal='center', vertical='center')
                ws['D7'].value = 'TS'
                ws['D7'].font = line_font
                ws['D7'].border = all_border_thin
                ws['D7'].alignment = Alignment(horizontal='center', vertical='center')
                ws['E7'].value = 'R'
                ws['E7'].font = line_font
                ws['E7'].border = all_border_thin
                ws['E7'].alignment = Alignment(horizontal='center', vertical='center')
                ws['F7'].value = 'Ro'
                ws['F7'].font = line_font
                ws['F7'].border = all_border_thin
                ws['F7'].alignment = Alignment(horizontal='center', vertical='center')
                ws['G7'].value = 'NF'
                ws['G7'].font = line_font
                ws['G7'].border = all_border_thin
                ws['G7'].alignment = Alignment(horizontal='center', vertical='center')
                ws['H7'].value = 'NS'
                ws['H7'].font = line_font
                ws['H7'].border = all_border_thin
                ws['H7'].alignment = Alignment(horizontal='center', vertical='center')
                ws['I7'].value = 'NC'
                ws['I7'].font = line_font
                ws['I7'].border = all_border_thin
                ws['I7'].alignment = Alignment(horizontal='center', vertical='center')
                ws['J7'].value = 'XNC-Kho'
                ws['J7'].font = line_font
                ws['J7'].border = all_border_thin
                ws['J7'].alignment = Alignment(horizontal='center', vertical='center')
                ws['K7'].value = 'RN'
                ws['K7'].font = line_font
                ws['K7'].border = all_border_thin
                ws['K7'].alignment = Alignment(horizontal='center', vertical='center')
                ws['L7'].value = 'CT'
                ws['L7'].font = line_font
                ws['L7'].border = all_border_thin
                ws['L7'].alignment = Alignment(horizontal='center', vertical='center')
                ws['M7'].value = 'Le'
                ws['M7'].font = line_font
                ws['M7'].border = all_border_thin
                ws['M7'].alignment = Alignment(horizontal='center', vertical='center')
                ws.merge_cells(start_row=9, start_column=1, end_row=9, end_column=16)
                ws.merge_cells(start_row=10, start_column=1, end_row=11, end_column=1)
                ws.merge_cells(start_row=10, start_column=2, end_row=11, end_column=2)
                ws.merge_cells(start_row=10, start_column=3, end_row=10, end_column=4)
                ws.merge_cells(start_row=10, start_column=5, end_row=10, end_column=6)
                ws.merge_cells(start_row=10, start_column=7, end_row=10, end_column=8)
                ws['A10'].value = 'Ngày'
                ws['A10'].font = line_font
                ws['A10'].border = all_border_thin
                ws['A10'].alignment = Alignment(horizontal='center', vertical='center')
                ws['B10'].value = 'Thứ'
                ws['B10'].font = line_font
                ws['B10'].border = all_border_thin
                ws['B10'].alignment = Alignment(horizontal='center', vertical='center')
                ws['C10'].value = '1'
                ws['C10'].font = line_font
                ws['C10'].border = all_border_thin
                ws['C10'].alignment = Alignment(horizontal='center', vertical='center')
                ws['C11'].value = 'Vào'
                ws['C11'].font = line_font
                ws['C11'].border = all_border_thin
                ws['C11'].alignment = Alignment(horizontal='center', vertical='center')
                ws['D11'].value = 'Ra'
                ws['D11'].font = line_font
                ws['D11'].border = all_border_thin
                ws['D11'].alignment = Alignment(horizontal='center', vertical='center')
                ws['E10'].value = '2'
                ws['E10'].font = line_font
                ws['E10'].border = all_border_thin
                ws['E10'].alignment = Alignment(horizontal='center', vertical='center')
                ws['E11'].value = 'Vào'
                ws['E11'].font = line_font
                ws['E11'].border = all_border_thin
                ws['E11'].alignment = Alignment(horizontal='center', vertical='center')
                ws['F11'].value = 'Ra'
                ws['F11'].font = line_font
                ws['F11'].border = all_border_thin
                ws['F11'].alignment = Alignment(horizontal='center', vertical='center')
                ws['G10'].value = '2'
                ws['G10'].font = line_font
                ws['G10'].border = all_border_thin
                ws['G10'].alignment = Alignment(horizontal='center', vertical='center')
                ws['G11'].value = 'Vào'
                ws['G11'].font = line_font
                ws['G11'].border = all_border_thin
                ws['G11'].alignment = Alignment(horizontal='center', vertical='center')
                ws['H11'].value = 'Ra'
                ws['H11'].font = line_font
                ws['H11'].border = all_border_thin
                ws['H11'].alignment = Alignment(horizontal='center', vertical='center')
                ws['I10'].value = 'Trễ'
                ws['I10'].font = line_font
                ws['I10'].border = all_border_thin
                ws['I10'].alignment = Alignment(horizontal='center', vertical='center')
                ws['J10'].value = 'Sớm'
                ws['J10'].font = line_font
                ws['J10'].border = all_border_thin
                ws['J10'].alignment = Alignment(horizontal='center', vertical='center')
                ws['K10'].value = 'Công'
                ws['K10'].font = line_font
                ws['K10'].border = all_border_thin
                ws['K10'].alignment = Alignment(horizontal='center', vertical='center')
                ws['L10'].value = 'T.Giờ'
                ws['L10'].font = line_font
                ws['L10'].border = all_border_thin
                ws['L10'].alignment = Alignment(horizontal='center', vertical='center')
                ws['M10'].value = 'T.Ca1'
                ws['M10'].font = line_font
                ws['M10'].border = all_border_thin
                ws['M10'].alignment = Alignment(horizontal='center', vertical='center')
                ws['N10'].value = 'T.Ca2'
                ws['N10'].font = line_font
                ws['N10'].border = all_border_thin
                ws['N10'].alignment = Alignment(horizontal='center', vertical='center')
                ws['O10'].value = 'T.Ca3'
                ws['O10'].font = line_font
                ws['O10'].border = all_border_thin
                ws['O10'].alignment = Alignment(horizontal='center', vertical='center')
                ws['P10'].value = 'Nơi làm việc'
                ws['P10'].font = line_font
                ws['P10'].border = all_border_thin
                ws['P10'].alignment = Alignment(horizontal='center', vertical='center')
                for atten in attendance:
                    wd = date.weekday(atten.date_attendance)
                    if atten.check_in:
                        check_in = atten.check_in.astimezone(tz)
                    if atten.check_out:
                        check_out = atten.check_out.astimezone(tz)
                    rec_data.append({
                        'date': atten.date_attendance.strftime("%d/%m/%Y"),
                        'sort': days[wd],
                        'check_in': check_in.strftime("%H:%M"),
                        'check_out': check_out.strftime("%H:%M") if atten.check_out else '-'
                    })
                print(rec_data, '_______________')
                key_col_list = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]
                key_list = [
                    'date',
                    'sort',
                    'check_in',
                    'check_out',
                ]
                row = 12
                for line_data in rec_data:
                    for col, k in zip(key_col_list, key_list):
                        cell = ws.cell(row, col)
                        cell.value = line_data[k]
                        cell.font = line_font
                        cell.border = all_border_thin
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    row += 1
                fp = BytesIO()
                wb.save(fp)
                fp.seek(0)
                report = base64.encodebytes((fp.read()))
                fp.close()
                attachment = self.env['ir.attachment'].sudo().create({
                    'name': 'Báo cáo chấm công chi tiet',
                    # 'datas_fname': 'bao_cao_vat_tu_theo_khoa_hoc.xlsx',
                    'datas': report,
                    'res_model': 'temp.creation',
                    'public': True,
                })
            return {
                'name': 'Báo cáo chấm công',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                # 'view_type': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}
            }