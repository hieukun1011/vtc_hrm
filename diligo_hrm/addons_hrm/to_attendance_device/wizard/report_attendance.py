import base64
import logging
from calendar import monthrange
from datetime import date, datetime
from io import BytesIO

import pytz
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

from odoo import fields, models

_logger = logging.getLogger(__name__)

days = ['T2', 'T3', 'T4', 'T5', 'T6', 'T7', 'CN']


class ReportAttendance(models.TransientModel):
    _name = 'report.attendance'

    type = fields.Selection([('department', 'Department'), ('employee', 'Employee'), ('area', 'Area')], string='Type',
                            default="area")
    date_start = fields.Date(string='Date From', required=True,
                             default=lambda self: fields.Date.to_string(
                                 date.today().replace(month=date.today().month - 1, day=1)))
    date_end = fields.Date(string='Date To', required=True,
                           default=lambda self: fields.Date.to_string(
                               (datetime.now() + relativedelta(day=1, days=-1)).date()))
    attendance = fields.Many2one('hr.attendance', string='Attendance')
    employee_id = fields.Many2one('hr.employee', string='Employee')
    department_id = fields.Many2one('hr.department', string='Department')
    area_id = fields.Many2one('hr.channel', string='Area')
    # total_worktime = fields.Float('Total working')


    def report_attendance(self):
        simple_pharmacy_provider_attachment = self.env['ir.attachment'].sudo().browse(
            self.env.ref('to_attendance_device.report_attendance_template').id)
        decode = base64.b64decode(simple_pharmacy_provider_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        ws.column_dimensions[get_column_letter(4)].width = 5
        ws.column_dimensions[get_column_letter(1)].width = 5
        ws.column_dimensions[get_column_letter(2)].width = 20

        print(self.type)
        ws['B6'].value = 'Mã nhân viên:'
        ws['C6'].value = 'Họ tên'
        ws['C6'].fill = PatternFill(patternType='solid',
                                    fgColor='C991E6')
        ws['D1'].value = 'BẢNG CHẤM CÔNG THÁNG ' + str(self.date_end.month) + ' NĂM ' + str(self.date_end.year)

        thin = borders.Side(style='thin')
        all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
        line_font = Font(name='Times New Roman', size=12)
        ws['D1'].font = Font(name='Times New Roman', size=20)
        # ws['D1'].border = all_border_thin
        # ws['D1'].alignment = Alignment(horizontal='left', vertical='center')
        monthrange(self.date_end.year, self.date_end.month)
        key_col_list = []
        key_col_list_date = []
        key_col_list_sort = []
        for rec in range(1, (monthrange(self.date_end.year, self.date_end.month)[1]) + 1):
            key_col_list.append(rec + 3)
            key_col_list_date.append(rec)
            sort = date(self.date_end.year, self.date_end.month, rec)
            wd = date.weekday(sort)
            key_col_list_sort.append(days[wd])
        key_list = [
            'employee_code',
            'employee',
            'data_atten',
            'total_hour',
            'data_holyday',
            'data_leave',
        ]
        row = 8
        cn = []
        t7 = []

        for col, k in zip(key_col_list, key_col_list_sort):
            cell = ws.cell(6, col)
            cell.fill = PatternFill(patternType='solid',
                                    fgColor='C991E6')
            cell.value = k
            cell.font = line_font
            cell.border = all_border_thin
            cell.alignment = Alignment(horizontal='left', vertical='center')
            if cell.value == 'CN':
                cn.append(col)
                cell.fill = PatternFill(patternType='solid',
                                        fgColor='FBE254')
            if cell.value == 'T7':
                t7.append(col)
        for col, d in zip(key_col_list, key_col_list_date):
            cell = ws.cell(7, col)
            cell.fill = PatternFill(patternType='solid',
                                    fgColor='C991E6')
            cell.value = d
            cell.font = line_font
            cell.border = all_border_thin
            cell.alignment = Alignment(horizontal='left', vertical='center')
        total_worktime = 0
        for dt in key_col_list_date:
            if dt not in cn:
                total_worktime += 8
        print(total_worktime)
        ws['C5'].value = 'Tông: ' + str(round(total_worktime)) + ' giờ'
        if self.type == 'department':
            rec_data = []
            atten = self.env['hr.attendance'].search([('employee_id', 'in', self.department_id.member_ids.ids),
                                                             ('check_in', '>=', self.date_start),
                                                             ('check_in', '<=', self.date_end),
                                                             ('check_out', '>=', self.date_start),
                                                             ('check_out', '<=', self.date_end),
                                                             ('state', '=', 'validate')
                                                             ])
            equipment_data = self.env['hr.attendance'].read_group([('employee_id', 'in', self.department_id.member_ids.ids),
                                                             ('check_in', '>=', self.date_start),
                                                             ('check_in', '<=', self.date_end),
                                                             ('check_out', '>=', self.date_start),
                                                             ('check_out', '<=', self.date_end),
                                                             ('state', '=', 'validate')], ['employee_id'],
                                                                    ['employee_id'])

            for item in equipment_data:
                data = atten.search([('employee_id', '=', item['employee_id'][0])])
            holydays = self.env['resource.calendar.leaves'].sudo().search([('date_to', '>=', self.date_start),
                                                                           ('date_from', '<=', self.date_end),
                                                                           ('resource_id', '=', False)
                                                                           ])
            late_registration = self.env['hr.leave'].sudo().search([('date_to', '>=', self.date_start),
                                                        ('date_from', '<=', self.date_end),
                                                        ('holiday_status_id.code', '=', 'DM'),
                                                        ('employee_id', 'in', self.department_id.member_ids.ids)
                                                        ])


            leave = self.env['hr.leave'].sudo().search([('date_to', '>=', self.date_start),
                                                        ('date_from', '<=', self.date_end),
                                                        ('employee_id', 'in', self.department_id.member_ids.ids),
                                                        ('state', '=', 'validate')
                                                        ])
            for member in self.department_id.member_ids:
                tz = pytz.timezone(member.tz)
                data = {
                    'employee_code': member.employee_id or "-",
                    'employee': member.name or "-",
                    'total_hour': 0,
                }
                data_late = []

                for k in late_registration:
                    if k.request_date_from >= self.date_start and k.request_date_from <= self.date_end and k.employee_id == member:
                        data_late.append({
                            'date': k.request_date_from,
                            'reason': k.name,
                        })
                data['data_late'] = data_late
                data_checkin = []
                for rec in atten:
                    day_checkin = datetime.date(rec.check_in.astimezone(tz)) or datetime.date(
                        rec.check_out.astimezone(tz))
                    if rec.employee_id == member:
                        data_checkin.append({
                            'check_in': day_checkin.day,
                            'worked_hours_real': round(rec.worked_hours_real, 1),
                            'reason': rec.reason,
                        })
                        data['total_hour'] += round(rec.worked_hours_real, 1)
                data['data_atten'] = data_checkin
                data_holyday = []
                for day in holydays:
                    if datetime.date(day.date_to) > self.date_end:
                        for i in range(day.date_from.day, self.date_end.day + 1):
                            if (i + 3) not in cn:
                                data_holyday.append({
                                    'date': i,
                                    'code': 'NL',
                                    'total_holiday': 8,
                                })
                    elif datetime.date(day.date_from) < self.date_start:
                        for i in range(self.date_start.day, day.date_to.day + 1):
                            if (i + 3) not in cn:
                                data_holyday.append({
                                    'date': i,
                                    'code': 'NL',
                                    'total_holiday': 8,
                                })
                    else:
                        for i in range(day.date_from.day, day.date_to.day + 1):
                            if (i + 3) not in cn:
                                data_holyday.append({
                                    'date': i,
                                    'code': 'NL',
                                    'total_holiday': 8,
                                })
                data['data_holyday'] = data_holyday
                data_leave = []
                for l in leave:
                    if l.employee_id == member:
                        if l.number_of_days >= 1:
                            if datetime.date(l.date_to) > self.date_end:
                                for i in range(l.date_from.day, self.date_end.day + 1):
                                    if (i + 3) in t7:
                                        data_leave.append({
                                            'date': i,
                                            'code': l.holiday_status_id.code,
                                            'total_leave': 4,
                                        })
                                    elif (i + 3) not in cn:
                                        data_leave.append({
                                            'date': i,
                                            'code': l.holiday_status_id.code,
                                            'total_leave': 8,
                                        })
                            elif datetime.date(l.date_from) < self.date_start:
                                for i in range(self.date_start.day, l.date_to.day + 1):
                                    if (i + 3) in t7:
                                        data_leave.append({
                                            'date': i,
                                            'code': l.holiday_status_id.code,
                                            'total_leave': 4,
                                        })
                                    elif (i + 3) not in cn:
                                        data_leave.append({
                                            'date': i,
                                            'code': l.holiday_status_id.code,
                                            'total_leave': 8,
                                        })
                            else:
                                for i in range(l.date_from.day, l.date_to.day + 1):
                                    if (i + 3) in t7:
                                        data_leave.append({
                                            'date': i,
                                            'code': l.holiday_status_id.code,
                                            'total_leave': 4,
                                        })
                                    elif (i + 3) not in cn:
                                        data_leave.append({
                                            'date': i,
                                            'code': l.holiday_status_id.code,
                                            'total_leave': 8,
                                        })
                        else:
                            data_leave.append({
                                'date': datetime.date(l.date_from.astimezone(tz)).day,
                                'day': l.number_of_days,
                                'code': l.holiday_status_id.code,
                                'total_leave': l.number_of_days * 8,
                            })
                data['data_leave'] = data_leave
                rec_data.append(data)
            # print(rec_data)
            for line_data in rec_data:
                ws.cell(row, 1).value = row - 7
                for col, k in zip(key_col_list, key_list):
                    cell = ws.cell(row, col - 2)
                    for c in line_data['data_late']:
                        date_late = c.get('date').day
                        ws.cell(row, date_late + 3).comment = Comment(c.get('reason'), "Author")
                    if k == 'total_hour':
                        ws.cell(row, len(key_col_list_date) + 8).value = line_data[k]
                    elif k == 'data_holyday':
                        for holyday in line_data[k]:
                            ws.cell(row, holyday.get('date') + 3).value = holyday['code']
                            ws.cell(row, len(key_col_list_date) + 6).value = len(line_data[k]) * 8
                    elif k == 'data_leave':
                        total_leave = 0
                        total_not_leave = 0
                        for lea in line_data[k]:
                            if lea['code'] == 'NP':
                                total_leave += lea[
                                    'total_leave']  # round(lea['total_leave'], 0) str(lea['total_leave'])
                                ws.cell(row, lea.get('date') + 3).value = (int(lea['total_leave']))
                                ws.cell(row, lea.get('date') + 3).fill = PatternFill(patternType='solid',
                                                         fgColor='00F700')
                            else:
                                total_not_leave += lea['total_leave']
                                ws.cell(row, lea.get('date') + 3).value = (int(lea['total_leave']))
                                ws.cell(row, lea.get('date') + 3).fill = PatternFill(patternType='solid',
                                                                                     fgColor='02A8EA')

                            # ws.cell(row, lea.get('date') + 3).value = lea['code'] #total_leave
                            ws.cell(row, len(key_col_list_date) + 11).value = total_leave
                            ws.cell(row, len(key_col_list_date) + 12).value = total_not_leave
                    elif k == 'data_atten':
                        for atten in line_data[k]:
                            ws.cell(row, atten.get('check_in') + 3).value = atten['worked_hours_real'] or 0

                            ws.cell(row, atten.get('check_in') + 3).font = line_font
                            ws.cell(row, atten.get('check_in') + 3).border = all_border_thin
                            ws.cell(row, atten.get('check_in') + 3).alignment = Alignment(horizontal='left',
                                                                                          vertical='center')
                            if atten.get('reason'):
                                ws.cell(row, atten.get('check_in') + 3).comment = Comment(atten.get('reason') or '-',
                                                                                          "Author")
                    else:
                        cell.value = line_data[k]

                    ws.cell(row, len(key_col_list_date) + 9).value = (ws.cell(row,
                                                                              len(key_col_list_date) + 8).value or 0) + (
                                                                                 ws.cell(row,
                                                                                         len(key_col_list_date) + 11).value or 0) + (
                                                                                 ws.cell(row,
                                                                                         len(key_col_list_date) + 6).value or 0)
                    ws.cell(row, len(key_col_list_date) + 10).value = total_worktime - (
                                ws.cell(row, len(key_col_list_date) + 9).value or 0) - (ws.cell(row,
                                                                                                len(key_col_list_date) + 12).value or 0)
                    cell.font = line_font
                    cell.border = all_border_thin
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                row += 1
            for i in range(1, len(key_col_list_date) + 17):
                for a in range(1, row - 5):
                    if i in cn:
                        ws.cell(a + 5, i).fill = PatternFill(patternType='solid',
                                                             fgColor='FBE254')
                    if i == (len(key_col_list_date) + 4):
                        ws.merge_cells(start_row=6, start_column=i, end_row=6, end_column=i + 12)
                        ws.cell(6, i).value = 'Tổng'
                        ws.cell(6, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                        ws.cell(6, i).alignment = Alignment(horizontal='center', vertical='center')
                        ws.cell(7, i).value = 'Làm chủ nhật'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 5):
                        ws.cell(7, i).value = 'OT'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 6):
                        ws.cell(7, i).value = 'Nghỉ lễ '
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 7):
                        ws.cell(7, i).value = 'Làm việc online'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 8):
                        ws.cell(7, i).value = 'Giờ làm việc thực tế'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 9):
                        ws.cell(7, i).value = 'Tổng giờ công'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 10):
                        ws.cell(7, i).value = 'Số giờ thiếu'  # F7020A
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='F7020A')
                    elif i == (len(key_col_list_date) + 11):
                        ws.cell(7, i).value = 'Số giờ nghỉ phép'  # 00F701
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='00F701')
                    elif i == (len(key_col_list_date) + 12):
                        ws.cell(7, i).value = 'Số giờ nghỉ k lương'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='5BC4E8')
                    elif i == (len(key_col_list_date) + 13):
                        ws.cell(7, i).value = 'Phép tháng 03'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 14):
                        ws.cell(7, i).value = 'Phép tháng 04'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 15):
                        ws.cell(7, i).value = 'Giờ phép còn lại'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 16):
                        ws.cell(7, i).value = 'Ký nhận'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    ws.cell(a + 5, i).font = line_font
                    ws.cell(a + 5, i).border = all_border_thin
                    ws.cell(a + 5, i).alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(start_row=len(rec_data) + 10, start_column=4, end_row=len(rec_data) + 10, end_column=8)
            ws.merge_cells(start_row=len(rec_data) + 10, start_column=27, end_row=len(rec_data) + 10, end_column=31)
            ws.merge_cells(start_row=len(rec_data) + 14, start_column=4, end_row=len(rec_data) + 14, end_column=8)
            ws.merge_cells(start_row=len(rec_data) + 14, start_column=27, end_row=len(rec_data) + 14, end_column=31)
            ws.cell(len(rec_data) + 10, 4).value = "Người Lập Bảng"
            ws.cell(len(rec_data) + 10, 4).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(len(rec_data) + 14, 4).value = self.env.user.name
            ws.cell(len(rec_data) + 14, 3).value = 'NL: Nghỉ lễ'
            ws.cell(len(rec_data) + 15, 3).value = 'Nghỉ không lương'
            ws.cell(len(rec_data) + 15, 3).fill = PatternFill(patternType='solid',
                                                         fgColor='5BC4E8')
            ws.cell(len(rec_data) + 16, 3).value = 'Nghỉ phép'
            ws.cell(len(rec_data) + 16, 3).fill = PatternFill(patternType='solid',
                                                         fgColor='00F701')
            ws.cell(len(rec_data) + 17, 3).value = 'Làm ở nhà'
            ws.cell(len(rec_data) + 17, 3).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
            ws.cell(len(rec_data) + 14, 4).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(len(rec_data) + 10, 27).value = "Trưởng Bộ Phận"
            ws.cell(len(rec_data) + 10, 27).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(len(rec_data) + 14, 27).value = "Nguyễn Hạnh Chung"
            ws.cell(len(rec_data) + 14, 27).alignment = Alignment(horizontal='center', vertical='center')
            fp = BytesIO()
            wb.save(fp)
            fp.seek(0)
            report = base64.encodebytes((fp.read()))
            fp.close()
            attachment = self.env['ir.attachment'].sudo().create({
                'name': 'Báo cáo chấm công',
                # 'datas_fname': 'bao_cao_vat_tu_theo_khoa_hoc.xlsx',
                'datas': report,
                'res_model': 'temp.creation',
                'public': True,
            })
            return {
                'name': 'Báo cáo chấm công',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                # 'view_type': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}
            }
        elif self.type == 'area':
            rec_data = []
            # men. = for item in self.area
            atten = self.env['hr.attendance'].search([('employee_id', 'in', self.area_id.member_ids.ids),
                                                      ('check_in', '>=', self.date_start),
                                                      ('check_in', '<=', self.date_end),
                                                      ('check_out', '>=', self.date_start),
                                                      ('check_out', '<=', self.date_end),
                                                      ('state', '=', 'validate')
                                                      ])
            equipment_data = self.env['hr.attendance'].read_group(
                [('employee_id', 'in', self.area_id.member_ids.ids),
                 ('check_in', '>=', self.date_start),
                 ('check_in', '<=', self.date_end),
                 ('check_out', '>=', self.date_start),
                 ('check_out', '<=', self.date_end),
                 ('state', '=', 'validate')], ['employee_id'],
                ['employee_id'])

            for item in equipment_data:
                data = atten.search([('employee_id', '=', item['employee_id'][0])])
            holydays = self.env['resource.calendar.leaves'].sudo().search([('date_to', '>=', self.date_start),
                                                                           ('date_from', '<=', self.date_end),
                                                                           ('resource_id', '=', False)
                                                                           ])
            late_registration = self.env['hr.leave'].sudo().search([('date_to', '>=', self.date_start),
                                                                    ('date_from', '<=', self.date_end),
                                                                    ('holiday_status_id.code', '=', 'DM'),
                                                                    ('employee_id', 'in',
                                                                     self.area_id.member_ids.ids)
                                                                    ])

            leave = self.env['hr.leave'].sudo().search([('date_to', '>=', self.date_start),
                                                        ('date_from', '<=', self.date_end),
                                                        ('employee_id', 'in', self.area_id.member_ids.ids),
                                                        ('state', '=', 'validate')
                                                        ])
            for member in self.area_id.member_ids:
                tz = pytz.timezone(member.tz)
                data = {
                    'employee_code': member.employee_id or "-",
                    'employee': member.name or "-",
                    'total_hour': 0,
                }
                data_late = []

                for k in late_registration:
                    if k.request_date_from >= self.date_start and k.request_date_from <= self.date_end and k.employee_id == member:
                        data_late.append({
                            'date': k.request_date_from,
                            'reason': k.name,
                        })
                data['data_late'] = data_late
                data_checkin = []
                for rec in atten:
                    day_checkin = datetime.date(rec.check_in.astimezone(tz)) or datetime.date(
                        rec.check_out.astimezone(tz))
                    if rec.employee_id == member:
                        data_checkin.append({
                            'check_in': day_checkin.day,
                            'worked_hours_real': round(rec.worked_hours_real, 1),
                            'reason': rec.reason,
                        })
                        data['total_hour'] += round(rec.worked_hours_real, 1)
                data['data_atten'] = data_checkin
                data_holyday = []
                for day in holydays:
                    if datetime.date(day.date_to) > self.date_end:
                        for i in range(day.date_from.day, self.date_end.day + 1):
                            if (i + 3) not in cn:
                                data_holyday.append({
                                    'date': i,
                                    'code': 'NL',
                                    'total_holiday': 8,
                                })
                    elif datetime.date(day.date_from) < self.date_start:
                        for i in range(self.date_start.day, day.date_to.day + 1):
                            if (i + 3) not in cn:
                                data_holyday.append({
                                    'date': i,
                                    'code': 'NL',
                                    'total_holiday': 8,
                                })
                    else:
                        for i in range(day.date_from.day, day.date_to.day + 1):
                            if (i + 3) not in cn:
                                data_holyday.append({
                                    'date': i,
                                    'code': 'NL',
                                    'total_holiday': 8,
                                })
                data['data_holyday'] = data_holyday
                data_leave = []
                for l in leave:
                    if l.employee_id == member:
                        if l.number_of_days >= 1:
                            if datetime.date(l.date_to) > self.date_end:
                                for i in range(l.date_from.day, self.date_end.day + 1):
                                    if (i + 3) in t7:
                                        data_leave.append({
                                            'date': i,
                                            'code': l.holiday_status_id.code,
                                            'total_leave': 4,
                                        })
                                    elif (i + 3) not in cn:
                                        data_leave.append({
                                            'date': i,
                                            'code': l.holiday_status_id.code,
                                            'total_leave': 8,
                                        })
                            elif datetime.date(l.date_from) < self.date_start:
                                for i in range(self.date_start.day, l.date_to.day + 1):
                                    if (i + 3) in t7:
                                        data_leave.append({
                                            'date': i,
                                            'code': l.holiday_status_id.code,
                                            'total_leave': 4,
                                        })
                                    elif (i + 3) not in cn:
                                        data_leave.append({
                                            'date': i,
                                            'code': l.holiday_status_id.code,
                                            'total_leave': 8,
                                        })
                            else:
                                for i in range(l.date_from.day, l.date_to.day + 1):
                                    if (i + 3) in t7:
                                        data_leave.append({
                                            'date': i,
                                            'code': l.holiday_status_id.code,
                                            'total_leave': 4,
                                        })
                                    elif (i + 3) not in cn:
                                        data_leave.append({
                                            'date': i,
                                            'code': l.holiday_status_id.code,
                                            'total_leave': 8,
                                        })
                        else:
                            data_leave.append({
                                'date': datetime.date(l.date_from.astimezone(tz)).day,
                                'day': l.number_of_days,
                                'code': l.holiday_status_id.code,
                                'total_leave': l.number_of_days * 8,
                            })
                data['data_leave'] = data_leave
                rec_data.append(data)
            # print(rec_data)
            for line_data in rec_data:
                ws.cell(row, 1).value = row - 7
                for col, k in zip(key_col_list, key_list):
                    cell = ws.cell(row, col - 2)
                    for c in line_data['data_late']:
                        date_late = c.get('date').day
                        ws.cell(row, date_late + 3).comment = Comment(c.get('reason'), "Author")
                    if k == 'total_hour':
                        ws.cell(row, len(key_col_list_date) + 8).value = line_data[k]
                    elif k == 'data_holyday':
                        for holyday in line_data[k]:
                            ws.cell(row, holyday.get('date') + 3).value = holyday['code']
                            ws.cell(row, len(key_col_list_date) + 6).value = len(line_data[k]) * 8
                    elif k == 'data_leave':
                        total_leave = 0
                        total_not_leave = 0
                        for lea in line_data[k]:
                            if lea['code'] == 'NP':
                                total_leave += lea[
                                    'total_leave']  # round(lea['total_leave'], 0) str(lea['total_leave'])
                                ws.cell(row, lea.get('date') + 3).value = (int(lea['total_leave']))
                                ws.cell(row, lea.get('date') + 3).fill = PatternFill(patternType='solid',
                                                                                     fgColor='00F700')
                            else:
                                total_not_leave += lea['total_leave']
                                ws.cell(row, lea.get('date') + 3).value = (int(lea['total_leave']))
                                ws.cell(row, lea.get('date') + 3).fill = PatternFill(patternType='solid',
                                                                                     fgColor='02A8EA')

                            # ws.cell(row, lea.get('date') + 3).value = lea['code'] #total_leave
                            ws.cell(row, len(key_col_list_date) + 11).value = total_leave
                            ws.cell(row, len(key_col_list_date) + 12).value = total_not_leave
                    elif k == 'data_atten':
                        for atten in line_data[k]:
                            ws.cell(row, atten.get('check_in') + 3).value = atten['worked_hours_real'] or 0

                            ws.cell(row, atten.get('check_in') + 3).font = line_font
                            ws.cell(row, atten.get('check_in') + 3).border = all_border_thin
                            ws.cell(row, atten.get('check_in') + 3).alignment = Alignment(horizontal='left',
                                                                                          vertical='center')
                            if atten.get('reason'):
                                ws.cell(row, atten.get('check_in') + 3).comment = Comment(atten.get('reason') or '-',
                                                                                          "Author")
                    else:
                        cell.value = line_data[k]

                    ws.cell(row, len(key_col_list_date) + 9).value = (ws.cell(row,
                                                                              len(key_col_list_date) + 8).value or 0) + (
                                                                             ws.cell(row,
                                                                                     len(key_col_list_date) + 11).value or 0) + (
                                                                             ws.cell(row,
                                                                                     len(key_col_list_date) + 6).value or 0)
                    ws.cell(row, len(key_col_list_date) + 10).value = total_worktime - (
                            ws.cell(row, len(key_col_list_date) + 9).value or 0) - (ws.cell(row,
                                                                                            len(key_col_list_date) + 12).value or 0)
                    cell.font = line_font
                    cell.border = all_border_thin
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                row += 1
            for i in range(1, len(key_col_list_date) + 17):
                for a in range(1, row - 5):
                    if i in cn:
                        ws.cell(a + 5, i).fill = PatternFill(patternType='solid',
                                                             fgColor='FBE254')
                    if i == (len(key_col_list_date) + 4):
                        ws.merge_cells(start_row=6, start_column=i, end_row=6, end_column=i + 12)
                        ws.cell(6, i).value = 'Tổng'
                        ws.cell(6, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                        ws.cell(6, i).alignment = Alignment(horizontal='center', vertical='center')
                        ws.cell(7, i).value = 'Làm chủ nhật'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 5):
                        ws.cell(7, i).value = 'OT'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 6):
                        ws.cell(7, i).value = 'Nghỉ lễ '
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 7):
                        ws.cell(7, i).value = 'Làm việc online'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 8):
                        ws.cell(7, i).value = 'Giờ làm việc thực tế'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 9):
                        ws.cell(7, i).value = 'Tổng giờ công'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 10):
                        ws.cell(7, i).value = 'Số giờ thiếu'  # F7020A
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='F7020A')
                    elif i == (len(key_col_list_date) + 11):
                        ws.cell(7, i).value = 'Số giờ nghỉ phép'  # 00F701
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='00F701')
                    elif i == (len(key_col_list_date) + 12):
                        ws.cell(7, i).value = 'Số giờ nghỉ k lương'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='5BC4E8')
                    elif i == (len(key_col_list_date) + 13):
                        ws.cell(7, i).value = 'Phép tháng 03'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 14):
                        ws.cell(7, i).value = 'Phép tháng 04'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 15):
                        ws.cell(7, i).value = 'Giờ phép còn lại'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 16):
                        ws.cell(7, i).value = 'Ký nhận'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    ws.cell(a + 5, i).font = line_font
                    ws.cell(a + 5, i).border = all_border_thin
                    ws.cell(a + 5, i).alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(start_row=len(rec_data) + 10, start_column=4, end_row=len(rec_data) + 10, end_column=8)
            ws.merge_cells(start_row=len(rec_data) + 10, start_column=27, end_row=len(rec_data) + 10, end_column=31)
            ws.merge_cells(start_row=len(rec_data) + 14, start_column=4, end_row=len(rec_data) + 14, end_column=8)
            ws.merge_cells(start_row=len(rec_data) + 14, start_column=27, end_row=len(rec_data) + 14, end_column=31)
            ws.cell(len(rec_data) + 10, 4).value = "Người Lập Bảng"
            ws.cell(len(rec_data) + 10, 4).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(len(rec_data) + 14, 4).value = self.env.user.name
            ws.cell(len(rec_data) + 14, 3).value = 'NL: Nghỉ lễ'
            ws.cell(len(rec_data) + 15, 3).value = 'Nghỉ không lương'
            ws.cell(len(rec_data) + 15, 3).fill = PatternFill(patternType='solid',
                                                              fgColor='5BC4E8')
            ws.cell(len(rec_data) + 16, 3).value = 'Nghỉ phép'
            ws.cell(len(rec_data) + 16, 3).fill = PatternFill(patternType='solid',
                                                              fgColor='00F701')
            ws.cell(len(rec_data) + 17, 3).value = 'Làm ở nhà'
            ws.cell(len(rec_data) + 17, 3).fill = PatternFill(patternType='solid',
                                                              fgColor='E2C3CA')
            ws.cell(len(rec_data) + 14, 4).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(len(rec_data) + 10, 27).value = "Trưởng Bộ Phận"
            ws.cell(len(rec_data) + 10, 27).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(len(rec_data) + 14, 27).value = "Nguyễn Hạnh Chung"
            ws.cell(len(rec_data) + 14, 27).alignment = Alignment(horizontal='center', vertical='center')
            fp = BytesIO()
            wb.save(fp)
            fp.seek(0)
            report = base64.encodebytes((fp.read()))
            fp.close()
            attachment = self.env['ir.attachment'].sudo().create({
                'name': 'Báo cáo chấm công',
                # 'datas_fname': 'bao_cao_vat_tu_theo_khoa_hoc.xlsx',
                'datas': report,
                'res_model': 'temp.creation',
                'public': True,
            })
            return {
                'name': 'Báo cáo chấm công',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                # 'view_type': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}
            }
        elif self.type == 'employee':
            rec_data = []
            atten = self.env['hr.attendance'].sudo().search([('employee_id', '=', self.employee_id.id),
                                                             ('check_in', '>=', self.date_start),
                                                             ('check_in', '<=', self.date_end),
                                                             ('check_out', '>=', self.date_start),
                                                             ('check_out', '<=', self.date_end),
                                                             ('state', '=', 'validate')
                                                             ])
            holydays = self.env['resource.calendar.leaves'].sudo().search([('date_to', '>=', self.date_start),
                                                                           ('date_from', '<=', self.date_end),
                                                                           ('resource_id', '=', False)
                                                                           ])
            late_registration = self.env['hr.leave'].sudo().search([('date_to', '>=', self.date_start),
                                                                    ('date_from', '<=', self.date_end),
                                                                    ('holiday_status_id.code', '=', 'DM'),
                                                                    ('employee_id', '=', self.employee_id.id)
                                                                    ])
            leave = self.env['hr.leave'].sudo().search([('date_to', '>=', self.date_start),
                                                        ('date_from', '<=', self.date_end),
                                                        ('employee_id', '=', self.employee_id.id),
                                                        ('state', '=', 'validate')
                                                        ])
            tz = pytz.timezone(self.employee_id.tz)
            data = {
                'employee_code': self.employee_id.employee_id or "-",
                'employee': self.employee_id.name or "-",
                'total_hour': 0,
            }
            data_late = []
            for k in late_registration:
                if k.request_date_from >= self.date_start and k.request_date_from <= self.date_end and k.employee_id == self.employee_id:
                    data_late.append({
                        'date': k.request_date_from,
                        'reason': k.name,
                    })
            data['data_late'] = data_late
            data_checkin = []
            for rec in atten:
                day_checkin = datetime.date(rec.check_in.astimezone(tz)) or datetime.date(
                    rec.check_out.astimezone(tz))
                if rec.employee_id == self.employee_id:
                    data_checkin.append({
                        'check_in': day_checkin.day,
                        'worked_hours_real': round(rec.worked_hours_real, 1),
                        'reason': rec.reason,
                    })
                    data['total_hour'] += round(rec.worked_hours_real, 1)

            data['data_atten'] = data_checkin
            data_holyday = []
            for day in holydays:
                if datetime.date(day.date_to) > self.date_end:
                    for i in range(day.date_from.day, self.date_end.day + 1):
                        if (i + 3) not in cn:
                            data_holyday.append({
                                'date': i,
                                'code': 'NL',
                                'total_holiday': 8,
                            })
                elif datetime.date(day.date_from) < self.date_start:
                    for i in range(self.date_start.day, day.date_to.day + 1):
                        if (i + 3) not in cn:
                            data_holyday.append({
                                'date': i,
                                'code': 'NL',
                                'total_holiday': 8,
                            })
                else:
                    for i in range(day.date_from.day, day.date_to.day + 1):
                        if (i + 3) not in cn:
                            data_holyday.append({
                                'date': i,
                                'code': 'NL',
                                'total_holiday': 8,
                            })
            data['data_holyday'] = data_holyday
            data_leave = []
            for l in leave:
                if l.employee_id == self.employee_id:
                    if l.number_of_days >= 1:
                        if datetime.date(l.date_to) > self.date_end:
                            for i in range(l.date_from.day, self.date_end.day + 1):
                                if (i + 3) in t7:
                                    data_leave.append({
                                        'date': i,
                                        'code': l.holiday_status_id.code,
                                        'total_leave': 4,
                                    })
                                elif (i + 3) not in cn:
                                    data_leave.append({
                                        'date': i,
                                        'code': l.holiday_status_id.code,
                                        'total_leave': 8,
                                    })
                        elif datetime.date(l.date_from) < self.date_start:
                            for i in range(self.date_start.day, l.date_to.day + 1):
                                if (i + 3) in t7:
                                    data_leave.append({
                                        'date': i,
                                        'code': l.holiday_status_id.code,
                                        'total_leave': 4,
                                    })
                                elif (i + 3) not in cn:
                                    data_leave.append({
                                        'date': i,
                                        'code': l.holiday_status_id.code,
                                        'total_leave': 8,
                                    })
                        else:
                            for i in range(l.date_from.day, l.date_to.day + 1):
                                if (i + 3) in t7:
                                    data_leave.append({
                                        'date': i,
                                        'code': l.holiday_status_id.code,
                                        'total_leave': 4,
                                    })
                                elif (i + 3) not in cn:
                                    data_leave.append({
                                        'date': i,
                                        'code': l.holiday_status_id.code,
                                        'total_leave': 8,
                                    })
                    else:
                        data_leave.append({
                            'date': datetime.date(l.date_from.astimezone(tz)).day,
                            'day': l.number_of_days,
                            'code': l.holiday_status_id.code,
                            'total_leave': l.number_of_days * 8,
                        })
            data['data_leave'] = data_leave
            rec_data.append(data)
            for line_data in rec_data:
                ws.cell(row, 1).value = row - 7
                for col, k in zip(key_col_list, key_list):
                    cell = ws.cell(row, col - 2)
                    for c in line_data['data_late']:
                        date_late = c.get('date').day
                        ws.cell(row, date_late + 3).comment = Comment(c.get('reason'), "Author")
                    if k == 'total_hour':
                        ws.cell(row, len(key_col_list_date) + 8).value = line_data[k]
                    elif k == 'data_holyday':
                        for holyday in line_data[k]:
                            ws.cell(row, holyday.get('date') + 3).value = holyday['code']
                            ws.cell(row, len(key_col_list_date) + 6).value = len(line_data[k]) * 8
                    elif k == 'data_leave':
                        total_leave = 0
                        total_not_leave = 0
                        for lea in line_data[k]:
                            if lea['code'] == 'NP':
                                total_leave += lea[
                                    'total_leave']  # round(lea['total_leave'], 0) str(lea['total_leave'])
                                ws.cell(row, lea.get('date') + 3).value = (int(lea['total_leave']))
                                ws.cell(row, lea.get('date') + 3).fill = PatternFill(patternType='solid',
                                                         fgColor='00F700')
                            else:
                                total_not_leave += lea['total_leave']
                                ws.cell(row, lea.get('date') + 3).value = (int(lea['total_leave']))
                                ws.cell(row, lea.get('date') + 3).fill = PatternFill(patternType='solid',
                                                                                     fgColor='02A8EA')

                            # ws.cell(row, lea.get('date') + 3).value = lea['code'] #total_leave
                            ws.cell(row, len(key_col_list_date) + 11).value = total_leave
                            ws.cell(row, len(key_col_list_date) + 12).value = total_not_leave
                    elif k == 'data_atten':
                        for atten in line_data[k]:
                            ws.cell(row, atten.get('check_in') + 3).value = atten['worked_hours_real'] or 0

                            ws.cell(row, atten.get('check_in') + 3).font = line_font
                            ws.cell(row, atten.get('check_in') + 3).border = all_border_thin
                            ws.cell(row, atten.get('check_in') + 3).alignment = Alignment(horizontal='left',
                                                                                          vertical='center')
                            if atten.get('reason'):
                                ws.cell(row, atten.get('check_in') + 3).comment = Comment(atten.get('reason') or '-',
                                                                                          "Author")
                    else:
                        cell.value = line_data[k]

                    ws.cell(row, len(key_col_list_date) + 9).value = (ws.cell(row,
                                                                              len(key_col_list_date) + 8).value or 0) + (
                                                                                 ws.cell(row,
                                                                                         len(key_col_list_date) + 11).value or 0) + (
                                                                                 ws.cell(row,
                                                                                         len(key_col_list_date) + 6).value or 0)
                    ws.cell(row, len(key_col_list_date) + 10).value = total_worktime - (
                                ws.cell(row, len(key_col_list_date) + 9).value or 0) - (ws.cell(row,
                                                                                                len(key_col_list_date) + 12).value or 0)
                    cell.font = line_font
                    cell.border = all_border_thin
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                row += 1
            for i in range(1, len(key_col_list_date) + 17):
                for a in range(1, row - 5):
                    if i in cn:
                        ws.cell(a + 5, i).fill = PatternFill(patternType='solid',
                                                             fgColor='FBE254')
                    if i == (len(key_col_list_date) + 4):
                        ws.merge_cells(start_row=6, start_column=i, end_row=6, end_column=i + 12)
                        ws.cell(6, i).value = 'Tổng'
                        ws.cell(6, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                        ws.cell(6, i).alignment = Alignment(horizontal='center', vertical='center')
                        ws.cell(7, i).value = 'Làm chủ nhật'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 5):
                        ws.cell(7, i).value = 'OT'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 6):
                        ws.cell(7, i).value = 'Nghỉ lễ '
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 7):
                        ws.cell(7, i).value = 'Làm việc online'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 8):
                        ws.cell(7, i).value = 'Giờ làm việc thực tế'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 9):
                        ws.cell(7, i).value = 'Tổng giờ công'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 10):
                        ws.cell(7, i).value = 'Số giờ thiếu'  # F7020A
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='F7020A')
                    elif i == (len(key_col_list_date) + 11):
                        ws.cell(7, i).value = 'Số giờ nghỉ phép'  # 00F701
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='00F701')
                    elif i == (len(key_col_list_date) + 12):
                        ws.cell(7, i).value = 'Số giờ nghỉ k lương'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='5BC4E8')
                    elif i == (len(key_col_list_date) + 13):
                        ws.cell(7, i).value = 'Phép tháng 03'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 14):
                        ws.cell(7, i).value = 'Phép tháng 04'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 15):
                        ws.cell(7, i).value = 'Giờ phép còn lại'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    elif i == (len(key_col_list_date) + 16):
                        ws.cell(7, i).value = 'Ký nhận'
                        ws.cell(7, i).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
                    ws.cell(a + 5, i).font = line_font
                    ws.cell(a + 5, i).border = all_border_thin
                    ws.cell(a + 5, i).alignment = Alignment(horizontal='center', vertical='center')
            # ws.cell(len(), 3).comment = Comment(c.get('reason'), "Author")
            ws.cell(len(rec_data) + 10, 4).value = "Người Lập Bảng"
            ws.cell(len(rec_data) + 10, 4).alignment = Alignment(horizontal='center', vertical='center')
            print(self.env.user)
            ws.cell(len(rec_data) + 14, 4).value = self.env.user.name
            ws.cell(len(rec_data) + 14, 3).value = 'NL: Nghỉ lễ'
            ws.cell(len(rec_data) + 15, 3).value = 'Nghỉ không lương'
            ws.cell(len(rec_data) + 15, 3).fill = PatternFill(patternType='solid',
                                                         fgColor='5BC4E8')
            ws.cell(len(rec_data) + 16, 3).value = 'Nghỉ phép'
            ws.cell(len(rec_data) + 16, 3).fill = PatternFill(patternType='solid',
                                                         fgColor='00F701')
            ws.cell(len(rec_data) + 17, 3).value = 'Làm ở nhà'
            ws.cell(len(rec_data) + 17, 3).fill = PatternFill(patternType='solid',
                                                         fgColor='E2C3CA')
            ws.cell(len(rec_data) + 14, 4).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(len(rec_data) + 10, 27).value = "Trưởng Bộ Phận"
            ws.cell(len(rec_data) + 10, 27).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(len(rec_data) + 14, 27).value = "Nguyễn Hạnh Chung"
            ws.cell(len(rec_data) + 14, 27).alignment = Alignment(horizontal='center', vertical='center')
            fp = BytesIO()
            wb.save(fp)
            fp.seek(0)
            report = base64.encodebytes((fp.read()))
            fp.close()
            attachment = self.env['ir.attachment'].sudo().create({
                'name': 'Báo cáo chấm công',
                # 'datas_fname': 'bao_cao_vat_tu_theo_khoa_hoc.xlsx',
                'datas': report,
                'res_model': 'temp.creation',
                'public': True,
            })
            return {
                'name': 'Báo cáo chấm công',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                # 'view_type': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}
            }
