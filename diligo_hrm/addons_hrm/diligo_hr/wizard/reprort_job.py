import base64
import logging
from calendar import monthrange
from datetime import *
from io import BytesIO
from time import strftime

import pytz
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill
from openpyxl.comments import Comment

from odoo import fields, models

_logger = logging.getLogger(__name__)

class ReportJob(models.TransientModel):
    _name = 'report.job'

    type = fields.Selection([('department', 'Department'), ('job', 'Job'), ('all', 'All')], string='Type',
                            default="department")
    date_start = fields.Datetime(string='Date From', required=True,
                             default=lambda self: fields.Date.to_string(
                                 date.today().replace(month=date.today().month - 1, day=1)))
    date_end = fields.Datetime(string='Date To', required=True,
                           default=lambda self: fields.Date.to_string(
                               (datetime.now() + relativedelta(day=1, days=-1)).date()))
    department_id = fields.Many2one('hr.department', string='Department')
    job_id = fields.Many2one('hr.job', string='Job')

    def report_job(self):
        simple_pharmacy_provider_attachment = self.env['ir.attachment'].sudo().browse(
            self.env.ref('diligo_hr.report_job_template3').id)
        decode = base64.b64decode(simple_pharmacy_provider_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        ws['B6'].value = 'PHÒNG BAN'
        ws['C6'].value = 'HỌ VÀ TÊN'
        ws['D6'].value = 'VỊ TRÍ ỨNG TUYỂN'
        ws['E6'].value = 'NGÀY ỨNG TUYỂN'
        ws['F6'].value = 'NGƯỜI PHỎNG VẤN'
        ws['G6'].value = 'NGÀY PHỎNG VẤN'
        ws['H6'].value = 'KẾT QUẢ'
        ws['I6'].value = 'NGÀY NHẬN VIỆC'
        ws['D1'].value = 'BẢNG BÁO CÁO CHI TIẾT THEO VỊ TRÍ TUYỂN DỤNG QUA WEB THÁNG ' + str(self.date_end.month) + ' NĂM ' + str(self.date_end.year)

        thin = borders.Side(style='thin')
        all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
        line_font = Font(name='Times New Roman', size=12)
        ws['D1'].font = Font(name='Times New Roman', size=20)
        monthrange(self.date_end.year, self.date_end.month)

        key_col_list = [2, 3, 4, 5, 6, 7, 8]
        key_list = [
            'department',
            'partner_name',
            'job_id',
            'create_date',
            'user_interview',
            'date_interview',
            'result',
            'joining_date',
        ]
        row = 8
        if self.type == 'department':
            rec_data = []
            department = self.env['hr.applicant'].sudo().search([
                ('department_id', '=', self.department_id.id),
            ])
            for item in department:
                data = {
                    'department': item.department_id.name or "-",
                    'partner_name': item.name or "-",
                    'job_id': item.job_id.name or "-",
                    'create_date': item.create_date or "-",
                    'user_interview': '-',
                    'date_interview': '-',
                    'result': '-',
                    'joining_date': '-',
                }
                if item.applicant_state:
                    user_interview = ''
                    if len(item.applicant_state[0].partner_ids) > 1:
                        for user in item.applicant_state[0].partner_ids:
                            user_interview += str(user.name + ', ')
                        data['user_interview'] = user_interview or '-',
                    else:
                        data['user_interview'] = item.applicant_state[0].partner_ids.name or '-',
                    data['date_interview'] = item.applicant_state[0].start or '-',
                    data['result'] = item.applicant_state[0].result or '-',
                if item.emp_id:
                    data['joining_date'] = item.emp_id.joining_date or '-',
                rec_data.append(data)
                # print(rec_data)

        elif self.type == 'job':
            rec_data = []
            jobs = self.env['hr.applicant'].sudo().search([
                ('job_id', '=', self.job_id.id),
                ('create_date', '>=', self.date_start),
                ('create_date', '<=', self.date_end),
            ])
            for item in jobs:
                data = {
                    'department': item.department_id.name or "-",
                    'partner_name': item.name or "-",
                    'job_id': item.job_id.name or "-",
                    'user_interview': '-',
                    'date_interview': '-',
                    'result': '-',
                    'joining_date': '-',
                }
                rec_data.append(data)
        else:
            rec_data = []
            al = self.env['hr.applicant'].sudo().search([
                ('date_open', '>=', self.date_start),
                ('date_open', '<=', self.date_end),
            ])
            for item in al:
                data = {
                    'department': item.department_id.name or "-",
                    'partner_name': item.name or "-",
                    'job_id': item.job_id.name or "-",
                    'create_date': item.create_date or "-",
                    'user_interview': '-',
                    'date_interview': '-',
                    'result': '-',
                    'joining_date': '-',
                }
                rec_data.append(data)

        for line_data in rec_data:
            ws.cell(row, 1).value = row - 7
            for col, k in zip(key_col_list, key_list):
                cell = ws.cell(row, col)
                cell.value = ''
                if type(line_data.get(k)) == tuple:
                    for i in line_data[k]:
                        cell.value += str(i)
                else:
                    cell.value = line_data.get(k)
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'Báo cáo chi tiết theo vị trí tuyển dụng qua web',
            # 'datas_fname': 'bao_cao_vat_tu_theo_khoa_hoc.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        return {
            'name': 'Báo cáo tiết theo vị trí tuyển dụng qua web',
            'type': 'ir.actions.act_window',
            'res_model': 'temp.wizard',
            'view_mode': 'form',
            # 'view_type': 'form',
            'target': 'inline',
            'view_id': self.env.ref('ms_templates.report_wizard').id,
            'context': {'attachment_id': attachment.id}
        }




