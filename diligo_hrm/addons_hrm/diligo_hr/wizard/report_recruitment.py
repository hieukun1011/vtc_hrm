import base64
import logging
import re
from datetime import date, datetime
from io import BytesIO

from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill
from openpyxl.comments import Comment

from odoo import fields, models

_logger = logging.getLogger(__name__)


class ReportRecruitment(models.TransientModel):
    _name = 'report.recruitment'

    type = fields.Selection([('department', 'Department'), ('job', 'Job'), ('all', 'All')], string='Type',
                            default="department")
    date_start = fields.Date(string='Date From', required=True,
                             default=lambda self: fields.Date.to_string(
                                 date.today().replace(month=date.today().month - 1, day=1)))
    date_end = fields.Date(string='Date To', required=True,
                           default=lambda self: fields.Date.to_string(
                               (datetime.now() + relativedelta(day=1, days=-1)).date()))
    job_id = fields.Many2one('hr.job', string='Job')
    department_id = fields.Many2one('hr.department', string='Department')

    def report_recruitment(self):
        def cleanhtml(raw_html):
            cleanr = re.compile('<.*?>')
            cleantext = re.sub(cleanr, '', str(raw_html))
            return cleantext

        def cleanhtmlspan(raw_html):
            cleanr = re.compile('</?span[^>]*>')
            cleantext = re.sub(cleanr, '', str(raw_html))
            return cleantext

        recruitment_report = self.env['ir.attachment'].browse(
            self.env.ref('diligo_hr.report_recruitment_template2').id)
        decode = base64.b64decode(recruitment_report.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        ws['A1'].value = "BÁO CÁO TUYỂN DỤNG TỔNG THẾ " + " THÁNG " + str(self.date_end.month) + " NĂM " + str(
            self.date_end.year)
        ws['A2'].value = "STT"
        ws['C2'].value = "CHI TIẾT"
        ws['B2'].value = "VỊ TRÍ CÔNG VIỆC"
        ws['D2'].value = "SL ỨNG TUYỂN"
        ws['E2'].value = "SL UV ĐÃ PV"
        ws['F2'].value = "SL UV ĐẠT"
        ws['G2'].value = "SL UV CHỜ XX"
        ws['H2'].value = "SL NS ĐANG CẦN"
        ws['I2'].value = "SL NS CÒN THIẾU"
        thin = borders.Side(style='thin')
        all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
        line_font = Font(name='Times New Roman', size=12)
        ws['A1'].font = Font(name='Times New Roman', size=20)
        key_col_list = [2, 3, 4, 5, 6, 7, 8, 9]
        key_list = [
            'position',
            'description',
            'count_applicant',
            'count_applicant_interview',
            'count_applicant_done',
            'count_applicant_xx',
            'count_personnel_need',
            'count_personnel_short',
        ]
        row = 3
        if self.type == 'department':
            rec_data = []
            department = self.env['hr.job'].sudo().search([
                ('department_id', '=', self.department_id.id)
            ])
            print(department)
            for item in department:
                application = self.env['hr.application'].sudo().search([('job_id', '=', item.id),
                                                                        ('create_date', '>=', self.date_start),
                                                                        ('create_date', '<=', self.date_end),
                                                                        ])
                stage = 0
                wait = 0
                for i in application:
                    for k in i.applicant_state:
                        if k.result == 'wait':
                            if k.stage_id.id == 2 or k.stage_id.id == 3:
                                wait += 1
                        if k.stage_id.id == 2:
                            stage += 1
                data = {
                    'position': item.name or '-',
                    'description': cleanhtml(cleanhtmlspan(item.description)) or '-',
                    'count_applicant': item.all_application_count or 0,
                    'count_applicant_interview': stage or 0,
                    'count_applicant_done': item.no_of_employee or 0,
                    'count_applicant_xx': wait or 0,
                    'count_personnel_need': item.total_payroll or 0,
                    'count_personnel_short': item.no_of_recruitment2 or 0,
                }
                rec_data.append(data)
        elif self.type == 'job':
            rec_data = []
            jobs = self.env['hr.job'].sudo().search([
                ('id', '=', self.job_id.id)
            ])
            for item in jobs:
                application = self.env['hr.application'].sudo().search([('job_id', '=', item.id),
                                                                        ('create_date', '>=', self.date_start),
                                                                        ('create_date', '<=', self.date_end),
                                                                        ])
                stage = 0
                wait = 0
                for i in application:
                    for k in i.applicant_state:
                        if k.result == 'wait':
                            if k.stage_id.id == 2 or k.stage_id.id == 3:
                                wait += 1
                        if k.stage_id.id == 2:
                            stage += 1
                data = {
                    'position': item.name or '-',
                    'description': cleanhtml(cleanhtmlspan(item.description)) or '-',
                    'count_applicant': item.all_application_count or 0,
                    'count_applicant_interview': stage or 0,
                    'count_applicant_done': item.no_of_employee or 0,
                    'count_applicant_xx': wait or 0,
                    'count_personnel_need': item.total_payroll or 0,
                    'count_personnel_short': item.no_of_recruitment2 or 0,
                }
                rec_data.append(data)
        else:
            rec_data = []
            job_all = self.env['hr.job'].sudo().search([])
            for item in job_all:
                application = self.env['hr.application'].sudo().search([('job_id', '=', item.id),
                                                                        ('create_date', '>=', self.date_start),
                                                                        ('create_date', '<=', self.date_end),
                                                                        ])
                stage = 0
                wait = 0
                for i in application:
                    for k in i.applicant_state:
                        if k.result == 'wait':
                            if k.stage_id.id == 2 or k.stage_id.id == 3:
                                wait += 1
                        if k.stage_id.id == 2:
                            stage += 1
                data = {
                    'position': item.name or '-',
                    'description': cleanhtml(cleanhtmlspan(item.description)) or '-',
                    'count_applicant': item.all_application_count or 0,
                    'count_applicant_interview': stage or 0,
                    'count_applicant_done': item.no_of_employee or 0,
                    'count_applicant_xx': wait or 0,
                    'count_personnel_need': item.total_payroll or 0,
                    'count_personnel_short': item.no_of_recruitment2 or 0,
                }
                rec_data.append(data)
        print(rec_data)
        for line_data in rec_data:
            ws.cell(row, 1).value = row - 2
            for col, k in zip(key_col_list, key_list):
                cell = ws.cell(row, col)
                cell.value = line_data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'Báo cáo tuyển dụng',
            # 'datas_fname': 'bao_cao_vat_tu_theo_khoa_hoc.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        return {
            'name': 'Báo cáo tuyển dụng',
            'type': 'ir.actions.act_window',
            'res_model': 'temp.wizard',
            'view_mode': 'form',
            # 'view_type': 'form',
            'target': 'inline',
            'view_id': self.env.ref('ms_templates.report_wizard').id,
            'context': {'attachment_id': attachment.id}
        }
